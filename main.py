import time
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

service = Service(ChromeDriverManager().install())

service.start()

driver = webdriver.Remote(service.service_url)

# driver.get('https://pizzadeliciousrugby.com/')
driver.get('https://www.deraj.co.uk/order-item/')


print('<<------------------START-------------------->>')

categories_labels = []
category_to_food = {}
food_dishes = []


# driver.quit()

# def getMenu():
#     menu = {}
#     dishes = {}
#     food = driver.find_elements(By.CLASS_NAME, 'productdes')
#     for category in food:
#         category_name = category.find_element(By.CLASS_NAME, 'MenuCategorystyles__CategoryTitle-sc-h1my9-6').text
#         category_items = []
#         print(category_name)
#         categories_labels.append(category_name)
#         time.sleep(5)
#         data = category.find_elements(By.CLASS_NAME, 'MenuItemstyles__MenuItemContent-sc-imari6-4')
#         if data == []:
#             print('No data found, Changing class name...')
#             data = category.find_element(By.CLASS_NAME, 'MenuCategorystyles__CategoryItems-sc-h1my9-11')
#         for item in data:
#             time.sleep(1.5)
#             category_items.append(item.text)
#             key = item.find_element(By.CLASS_NAME, 'MenuItemstyles__MenuItemTitle-sc-imari6-2')
#             if key.text == '':
#                 print('No key found, Changing to another class...')
#                 key = item.find_element(By.CLASS_NAME, 'MenuItemstyles__MenuItemTitleWrapper-sc-imari6-1')
#             discription = item.find_element(By.CLASS_NAME, 'MenuItemstyles__MenuItemDescription-sc-imari6-10')
#             if discription.text == '':
#                 print('No discription found, Using key as description...')
#                 discription = key
#             price = item.find_element(By.CLASS_NAME, 'PriceTagstyles__Price-sc-1j71n0t-1')
#             if price.text == '':
#                 print('No price found, Finding in another class...')
#                 price = item.find_element(By.CLASS_NAME, 'MenuItemstyles__MenuItemFooterWrapper-sc-imari6-11')
#             dishes[key.text] = [discription.text, price.text]
#             price_int = 'price.text'
#             price_int.replace('£','')
#             category_to_food[key.text] = [key.text, discription.text, price.text, category_name]
#             print(dishes[key.text])
#             food_dishes.append(key.text)
#         menu[category_name] = category_items
        
#         print('=========')
#     print(menu)
#     print('<<-------------------END------------------->>')
#     print(dishes)
    
def getCategories():
    ctgrs = driver.find_elements(By.CLASS_NAME, 'block-head')
    for ctgr in ctgrs:
        print(ctgr.text)
    
    
    
def getMenu():
    menu = driver.find_element(By.CLASS_NAME, 'productdes')
getMenu()


wb = load_workbook(filename="foods_bulk_format.xlsx")
ws = wb.active



def exportData():
    print('CATEGORIES>>>>>>\n')
    print(categories_labels)
    print('\n============================\n')
    print('MENU>>>>>>>>\n')
    index = 1
    for cat in category_to_food:
        
        label = food_dishes[index - 1]
        value = category_to_food[label]
        
        
        # FOOD WRITE OUT
        ws['A' + str(index + 1)] = str(value[0])
        ws['B' + str(index + 1)] = '1'
        ws['C' + str(index + 1)] = '0'
        ws['D' + str(index + 1)] = '0'
        ws['E' + str(index + 1)] = ''
        ws['F' + str(index + 1)] = str(value[2])
        ws['G' + str(index + 1)] = '0'
        ws['H' + str(index + 1)] = '0'
        ws['I' + str(index + 1)] = 'percent'
        ws['J' + str(index + 1)] = str(value[1])
        ws['K' + str(index + 1)] = '16:30:00'
        ws['L' + str(index + 1)] = '23:59:00'
        print(str(label) + '  ' + str(value[3]))
        index = index + 1
            
    wb.save("foods_bulk_format.xlsx")
        
        
        
# exportData()
    

